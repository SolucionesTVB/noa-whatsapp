import pandas as pd
from flask import Flask, render_template
from openpyxl import load_workbook

app = Flask(__name__)

@app.route("/tabla")
def mostrar_tabla():
    # Cargar el archivo Excel
    excel_path = "sugeseM_datos.xlsx"
    df = pd.read_excel(excel_path)
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    columnas_con_enlaces = ["Condiciones generales", "Solicitud de seguro", "Propuesta de seguro", "DERSA"]

    for col in columnas_con_enlaces:
        if col in df.columns:
            col_idx = df.columns.get_loc(col) + 1  # openpyxl usa 1-based indexing
            enlaces = []
            for i in range(len(df)):
                celda = ws.cell(row=i+2, column=col_idx)  # +2 porque openpyxl empieza en 1 y salteamos encabezado
                if celda.hyperlink:
                    url = celda.hyperlink.target
                    enlaces.append(f'<a href="{url}" target="_blank" class="boton-descarga">📄 Ver PDF</a>')
                else:
                    enlaces.append('<span class="no-disponible">— No disponible —</span>')
            df[col] = enlaces

    tabla_html = df.to_html(classes="tabla", escape=False, index=False)
    return render_template("tabla.html", tabla=tabla_html)
