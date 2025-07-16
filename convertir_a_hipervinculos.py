from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
# Ruta del archivo original
archivo_entrada = Path("sugeseM_datos.xlsx")
archivo_salida = Path("sugeseM_hipervinculos.xlsx")

# Columnas donde buscar posibles URLs como texto
columnas_objetivo = [
    "Condiciones generales",
    "Solicitud de seguro",
    "Propuesta de seguro",
    "DERSA"
]

# Abrimos el archivo original
wb = load_workbook(archivo_entrada)
ws = wb.active

# Identificamos el índice de cada columna objetivo
encabezados = {cell.value: idx for idx, cell in enumerate(ws[1]) if cell.value in columnas_objetivo}

for row in ws.iter_rows(min_row=2):
    for nombre_columna, col_idx in encabezados.items():
        celda = row[col_idx]
        valor = str(celda.value).strip()

        if valor.startswith("http://") or valor.startswith("https://"):
            celda.value = nombre_columna  # Texto que mostrará la celda
            celda.hyperlink = valor

# Guardamos en archivo nuevo
wb.save(archivo_salida)
print("✅ Hipervínculos aplicados exitosamente. Archivo guardado como:", archivo_salida)
wb.save("sugeseM_hipervinculos.xlsx")
