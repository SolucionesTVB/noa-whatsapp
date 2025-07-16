from openpyxl import load_workbook
from pathlib import Path

def extraer_hipervinculos(excel_path: Path, columnas_objetivo: list[str]) -> dict:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    enlaces_por_columna = {col: [] for col in columnas_objetivo}
    encabezados = {cell.value: idx for idx, cell in enumerate(ws[1])}

    for fila in ws.iter_rows(min_row=2):
        for columna in columnas_objetivo:
            col_idx = encabezados.get(columna)
            if col_idx is None:
                enlaces_por_columna[columna].append(
                    '<span class="no-disponible">— No disponible —</span>'
                )
                continue

            celda = fila[col_idx]
            enlace = celda.hyperlink.target if celda.hyperlink else None

            if enlace:
                ext = enlace.lower().split(".")[-1]
                if ext == "pdf":
                    texto = "📄 Ver PDF"
                    clase = "pdf"
                elif ext in {"doc", "docx"}:
                    texto = "📘 Ver Word"
                    clase = "word"
                elif ext in {"xls", "xlsx"}:
                    texto = "📊 Ver Excel"
                    clase = "excel"
                else:
                    texto = "🔗 Ver archivo"
                    clase = "otro"

                enlace_html = f'<a href="{enlace}" target="_blank" class="boton-enlace {clase}">{texto}</a>'
            else:
                enlace_html = '<span class="no-disponible">— No disponible —</span>'

            enlaces_por_columna[columna].append(enlace_html)

    return enlaces_por_columna
