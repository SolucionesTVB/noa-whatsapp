from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# Crear un nuevo libro y hoja
wb = Workbook()
ws = wb.active
ws.title = "Pólizas demo"

# Cabecera
ws.append(["Producto", "Condiciones generales", "Solicitud de seguro"])

# Fila 1: con dos hipervínculos
ws["A2"] = "Seguro Auto"
ws["B2"] = "Condiciones Auto"
ws["B2"].hyperlink = "https://www.example.com/pdf/condiciones_auto.pdf"
ws["C2"] = "Solicitud Auto"
ws["C2"].hyperlink = "https://www.example.com/pdf/solicitud_auto.pdf"

# Fila 2: con solo un hipervínculo
ws["A3"] = "Seguro Hogar"
ws["B3"] = "Condiciones Hogar"
ws["B3"].hyperlink = "https://www.example.com/pdf/condiciones_hogar.pdf"
ws["C3"] = ""  # Vacío

# Guardar el archivo
wb.save("demo_sugese.xlsx")
print("✅ Archivo demo_sugese.xlsx creado con hipervínculos reales.")
